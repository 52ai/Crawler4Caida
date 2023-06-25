# coding:utf-8
"""
create on May 6, 2023 By Wenyan YU
Email: ieeflsyu@outlook.com

Function:

所五一答题活动，需要评卷，200多号人，4套卷子
如果人工判的话，还挺费劲儿的，索性写个批卷的程序，大致流程如下：

1）把4套卷子的答案处理好，放到列表中；
2）读取文件夹中的6个excel文件，逐一逐行批改，然后将得分写到每行的后面


"""
import openpyxl
import os
import time
import re


answer_dic = {
    "A": ["D", "C", "C", "D", "C",
          "C", "C", "B", "C", "C",
          "B", "C", "D", "C", "ABCDE",
          "ABCDE", "ABCDE", "ABCDE", "ACDE", "ABCEF"],
    "B": ["A", "B", "B", "C", "D",
          "C", "D", "B", "C", "B",
          "B", "B", "A", "C", "BCDE",
          "BCDEF", "ABCDE", "ABDEF", "ABDE", "ABCDE"],
    "C": ["D", "C", "C", "C", "C",
          "B", "C", "C", "B", "C",
          "D", "C", "D", "C", "ABCD",
          "ABCEF", "ABCDF", "ABCDE", "ABCDE", "BCDE"],
    "D": ["C", "C", "A", "C", "B",
          "C", "C", "D", "C", "B",
          "B", "B", "D", "D", "BCDE",
          "BCDEF", "ABCDE", "ABDEF", "ABDE", "ABCDE"],

}


def judge(answer_file):
    """
    读取试卷，逐行评卷
    :param answer_file:
    :return:
    """
    print(answer_file)
    answer_str = answer_file.strip().split("/")[-1][0]
    print("当前答题的试卷：", answer_str)
    print("标准答案:", answer_dic[answer_str])
    work_book = openpyxl.load_workbook(answer_file)
    work_sheet = work_book.worksheets[0]
    rows_cnt = 0
    for row in work_sheet.rows:
        if rows_cnt <= 1:
            rows_cnt += 1
            continue
        name = row[0].value
        print("答题人：", name)
        cols_cnt = 0
        scores = 0
        for item_answer in row[5:25]:
            gain_answer = re.findall('[a-zA-Z]', item_answer.value)
            gain_answer = "".join(gain_answer)
            # print("试卷选择：", gain_answer)
            correct_answer = answer_dic[answer_str][cols_cnt]
            cols_cnt += 1
            # print("正确答案：", correct_answer)
            if gain_answer == correct_answer:
                scores += 5
        rows_cnt += 1
        print("最终得分：", scores)
        work_sheet.cell(rows_cnt, 27, scores)
    work_sheet.cell(2, 27, "最终得分")
    work_book.save(filename=answer_file)


if __name__ == "__main__":
    time_start = time.time()  # 记录启动时间
    file_path = []
    for root, dirs, files in os.walk("../000LocalData/GH202351QA_Judge/"):
        for file_item in files:
            file_path.append(os.path.join(root, file_item))
    for item in file_path:
        judge(item)
    print("=>Scripts Finish, Time Consuming:", (time.time() - time_start), "S")
