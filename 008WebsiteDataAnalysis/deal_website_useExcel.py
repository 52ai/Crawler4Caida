# coding:utf-8
"""
create on Mar 17,2019 by Wayne
"""
import openpyxl
import csv
import time

self_top_domain = []


def deal_domain_str(domain):
    """
    遍历self_top_domain，匹配domain，是否存在该顶级域
    如果存在则返回
    :param domain:
    :return: re_domain
    """
    re_domain = []
    if domain is None:  # 如果domain 为空，则直接返回空
        return re_domain
    for top_domain_item in self_top_domain:
        # print(top_domain_item.lower())
        if domain.find(top_domain_item.lower()) != -1:
            re_domain.append(top_domain_item)
    return re_domain


if __name__ == "__main__":
    # 获取自主顶级域名列表
    domain_file = "../000LocalData/domain.csv"
    websiteList_file = "../000LocalData/websiteList_deal.xlsx"
    csv_open = open(domain_file, 'r', encoding='utf-8-sig')
    for line in csv_open.readlines():
        line = line.strip().split(",")
        self_top_domain.extend(line)
    print(self_top_domain)

    workbook = openpyxl.load_workbook(websiteList_file)
    worksheet = workbook.worksheets[0]
    # for row in worksheet.rows:
    #     for cell in row:
    #         print(cell.value, end="")
    #     print()
    rows_cnt = 1  # 行计数，初始化为1
    for cell in list(worksheet.columns)[7]:
        if rows_cnt == 1:
            rows_cnt += 1
            continue
        domain_str = cell.value
        print(cell.value)
        top_domain = deal_domain_str(domain_str)
        print(top_domain)
        if len(top_domain) == 0:
            worksheet.cell(rows_cnt, 1, "否")
        else:
            worksheet.cell(rows_cnt, 1, "是")
            worksheet.cell(rows_cnt, 2, str(top_domain))
        rows_cnt += 1
    edit_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(time.time()))
    worksheet.cell(1, 1, "是否使用自主顶级域(%s)" % edit_time)
    workbook.save(filename=websiteList_file)