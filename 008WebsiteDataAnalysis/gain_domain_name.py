# coding:utf-8
"""
create on June 17, 2019 by Wayne
程序功能描述:
1.提取网站的二级域名；2.若存在com.cn,edu.cn,gov.cn,net.cn,org.cn等自主域名，则取到三级域名

基本思路如下：
首先获取到给定网址的netloc，亦即主机地址；
结合DNS根区文件中的所有顶级域名，对主机地址进行处理，按要求提取制定的域名（二级或者三级）
"""

import openpyxl
import time
from urllib.parse import urlparse
topHostPostfix = []  # 读取并存储root_zone_database中的所有顶级域名


def deal_domain_str(domain):
    """
    按要求处理网站地址
    :param domain:
    :return: re_domain
    """
    re_domain = ""
    if domain is None:  # 如果domain为空，则直接返回空
        return re_domain
    # 第一轮处理，获取网站的netloc，并去掉www前缀
    res = urlparse(domain)
    res_loc = res.netloc
    print(res)
    res_loc = res_loc.split('.')
    if res_loc[0] == 'www':
        del(res_loc[0])
    print(res_loc)
    # 第二轮处理，如果第一轮处理结果的split('.')之后的列表长度大于2，说明存在(...)x.gov.cn这样的情况或者(...)x.y.com的情况
    # 针对上述两种情况分别进行处理
    if len(res_loc) > 2:  # 进行第二轮处理
        res_loc = res_loc[::-1]
        print(res_loc)
        # 将逆置后的domain列表与topHostPostfix做匹配
        # 若匹配成功，则下一位；若匹配不成功，则保留当前项，并删除后面的项
        flag_cnt = 0  # 记录匹配不成功的个数
        for i in range(0, len(res_loc)):
            if res_loc[i] not in topHostPostfix:
                # print("Not in!")
                flag_cnt += 1
                if flag_cnt > 1:
                    res_loc[i] = ""  # 除第一次不匹配外，其余不匹配项均设为空，代表删除
        res_loc = res_loc[::-1]  # 将逆置的顺序调整过来
        res_loc = [item for item in res_loc if item != ""]  # 去除列表中的空值
    for item in res_loc:
        re_domain = re_domain + item + "."
    re_domain = re_domain[:-1]  # 去掉最后一个字符
    # for top_domain_item in self_top_domain:
    #     if domain.find(top_domain_item.lower()) != -1:
    #         re_domain.append(top_domain_item)
    return re_domain


if __name__ == "__main__":
    # 读取根区文件中所有的顶级域名
    root_zone_database_file = "../000LocalData/root_zone_database.xlsx"
    workbook_domain = openpyxl.load_workbook(root_zone_database_file)
    worksheet_domain = workbook_domain.worksheets[0]
    for cell in list(worksheet_domain.columns)[0]:
        top_domain_name = cell.value
        top_domain_name = top_domain_name[1:]  # 删除第一个字符
        topHostPostfix.append(top_domain_name)
    print(topHostPostfix)
    # 读取网站列表
    websiteList_file = "../000LocalData/t_website20190522_new.xlsx"
    workbook = openpyxl.load_workbook(websiteList_file)
    worksheet = workbook.worksheets[0]
    rows_cnt = 1  # 行计数
    for cell in list(worksheet.columns)[1]:
        # 跳过第一行
        if rows_cnt == 1:
            rows_cnt += 1
            continue
        domain_str = cell.value
        print(cell.value)
        top_domain = deal_domain_str(domain_str)
        print(top_domain)
        worksheet.cell(rows_cnt, 3, str(top_domain))
        rows_cnt += 1
    workbook.save(filename=websiteList_file)
